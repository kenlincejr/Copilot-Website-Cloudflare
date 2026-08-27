#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build-audit-workbook.py
=======================

Regenerates ``copilot-adoption-audit-workbook.xlsx`` at the repository root.

The workbook is the deliverable described in Section 10 of
``copilot-adoption-audit.html`` and specified by
``specs/copilot-adoption-audit-buildout.spec.md`` (Part A). It is committed as a
builder rather than maintained as a binary so that the file is reproducible.

Design rules that must not be broken (spec A.1 / A.2):

* Columns are detected by **header text**, never by position. Every read of the
  pasted export goes through ``MATCH`` on ``Setup`` and ``INDEX`` everywhere
  else. The export's column order varies with the columns the admin switched on.
* No ``XLOOKUP`` / ``XMATCH`` / ``SORT`` / ``FILTER`` / ``UNIQUE`` / ``SEQUENCE``
  anywhere. ``PERCENTILE``, not ``PERCENTILE.INC``.
* ``MEDIAN`` runs over helper columns on the hidden ``Calc`` sheet. No
  ``MEDIAN(IF(...))`` array formulas.
* Every computed cell is a formula. Nothing is precomputed in Python.
* Arial throughout. No currency figure anywhere in the workbook.

Usage::

    python tools/build-audit-workbook.py [output.xlsx]
    python scripts/recalc.py copilot-adoption-audit-workbook.xlsx 120
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.chart.marker import Marker
from openpyxl.comments import Comment
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

# --------------------------------------------------------------------------
# Constants
# --------------------------------------------------------------------------

LAST_ROW = 1001          # last worksheet row of the 1,000-row capacity
CAPACITY = LAST_ROW - 1  # 1,000 users

PASTE = "'Paste Export'!$A$2:$BZ$1001"
HDR = "'Paste Export'!$A$1:$BZ$1"

NAVY = "003057"
TEAL = "0F7C8A"
YELLOW = "FFFF00"
MUTED = "F3F4F6"
RULE = "D1D5DB"

F_BLUE = "FF0000FF"      # typed input values
F_GREEN = "FF008000"     # link to another sheet
F_RED = "FFC00000"       # warnings
F_GREY = "FF6B7280"
F_NAVY = "FF003057"

ARIAL = "Arial"

# The eighteen exact Microsoft header strings, in the order Microsoft ships
# them. Order is documentation only -- the workbook detects columns by text.
HEADERS = [
    "User name",
    "Display name",
    "Prompts submitted (any app)",
    "Copilot Chat (work) prompts submitted",
    "Copilot Chat (web) prompts submitted",
    "Active Days",
    "Last activity date (UTC)",
    "Last activity date of Teams Copilot (UTC)",
    "Last activity date of Word Copilot (UTC)",
    "Last activity date of Excel Copilot (UTC)",
    "Last activity date of PowerPoint Copilot (UTC)",
    "Last activity date of Outlook Copilot (UTC)",
    "Last activity date of OneNote Copilot (UTC)",
    "Last activity date of Loop Copilot (UTC)",
    "Last activity date of Copilot Chat (work) (UTC)",
    "Last activity date of Copilot Chat (web) (UTC)",
    "Last activity date of Microsoft 365 App (UTC)",
    "Last activity date of Microsoft Edge (UTC)",
]

# Three example rows: one heavy, one middling, one drifting.
EXAMPLES = [
    [
        "A1B2C3D4E5F60718", "9F8E7D6C5B4A3928", 612, 470, 142, 47, "2026-08-25",
        "2026-08-25", "2026-08-22", "2026-08-19", "2026-08-11", "2026-08-25",
        "2026-07-30", "2026-08-06", "2026-08-25", "2026-08-24", "2026-08-25",
        "2026-08-21",
    ],
    [
        "11223344556677AA", "BBCCDDEEFF001122", 96, 88, 8, 31, "2026-08-24",
        "2026-08-24", "2026-08-18", "", "", "2026-08-24", "", "",
        "2026-08-24", "", "2026-08-13", "",
    ],
    [
        "DEADBEEF01234567", "0123456789ABCDEF", 11, 6, 5, 3, "2026-06-12",
        "2026-06-12", "2026-05-28", "", "2026-05-21", "2026-06-02", "", "",
        "2026-06-12", "2026-06-11", "", "",
    ],
]

BANDS = ["Embedded", "Loyal but shallow", "Project-driven", "Drifting"]

BAND_RULE = {
    "Embedded": "Active days >= frequency line AND prompts >= depth line",
    "Loyal but shallow": "Active days >= frequency line AND prompts < depth line",
    "Project-driven": "Active days < frequency line AND prompts >= depth line",
    "Drifting": "Active days < frequency line AND prompts < depth line",
}

BAND_MEANING = {
    "Embedded": (
        "Frequent and deep. This is where the return is concentrated, and it is "
        "usually thinner than anyone expects. Check agent access before anyone "
        "proposes building agents."
    ),
    "Loyal but shallow": (
        "Present more often than the threshold, running a low rate per active "
        "day. Months of loyalty to a single use case. The highest conversion "
        "rate you will find: sell a second use case, not a mindset."
    ),
    "Project-driven": (
        "Rare visits, deep sessions. Capable and not yet embedded in anything "
        "weekly. Workflow embedding, and check their grounding -- this band "
        "skews to ungrounded chat."
    ),
    "Drifting": (
        "Fails both lines. Split it before you act on it: people who tried "
        "several surfaces and stopped are recoverable with use-case design; "
        "people who never started may be a licensing error."
    ),
}

# Column detection block: (label, exact header text, defined name)
DETECT = [
    ("User", "User name", "col_user"),
    ("Prompts (depth axis)", "Prompts submitted (any app)", "col_prompts"),
    ("Web prompts", "Copilot Chat (web) prompts submitted", "col_web"),
    ("Active days (frequency axis)", "Active Days", "col_days"),
    ("Last activity", "Last activity date (UTC)", "col_last"),
    ("Surface 1 - Teams", "Last activity date of Teams Copilot (UTC)", "col_s1"),
    ("Surface 2 - Word", "Last activity date of Word Copilot (UTC)", "col_s2"),
    ("Surface 3 - Excel", "Last activity date of Excel Copilot (UTC)", "col_s3"),
    ("Surface 4 - PowerPoint", "Last activity date of PowerPoint Copilot (UTC)", "col_s4"),
    ("Surface 5 - Outlook", "Last activity date of Outlook Copilot (UTC)", "col_s5"),
    ("Surface 6 - OneNote", "Last activity date of OneNote Copilot (UTC)", "col_s6"),
    ("Surface 7 - Loop", "Last activity date of Loop Copilot (UTC)", "col_s7"),
    ("Surface 8 - Microsoft 365 App", "Last activity date of Microsoft 365 App (UTC)", "col_s8"),
]
SURF_NAMES = [d[2] for d in DETECT[5:]]

TRAPS = [
    ("01", "Active Days is a chat measure, not an all-app measure",
     "Microsoft counts days the user submitted prompts to Copilot Chat. Call the "
     "frequency axis chat frequency, and cross-check the Drifting band against the "
     "per-app last-activity columns before calling anyone dormant."),
    ("02", "Not every Copilot action counts as a prompt",
     "Edit with Word counts toward Copilot Chat (work) prompts; Edit with Excel and "
     "Edit with PowerPoint do not. Teams Intelligent Recap, Interpreter and "
     "Facilitator count as active usage without generating a prompt. Prompt count "
     "is not a measure of effort."),
    ("03", "The denominator includes people who are no longer licensed",
     "The user-level table shows everyone who held a Copilot licence at any point "
     "in the past 180 days. Reconcile the row count against assigned licences "
     "under Billing > Licenses before computing any rate."),
    ("04", "The agent report and the usage report do not share a window",
     "Usage report: 7, 28, 90 or 180 days. Agents usage report: 7 or 30 days only. "
     "Never put an agent percentage and a usage percentage on one slide as though "
     "they covered the same period."),
    ("05", "The agent report changed what it counts",
     "The superseded report counted only agents your organisation built. The "
     "current one includes Microsoft and third-party agents and breaks usage down "
     "by creator type, so the same tenant looks far more agentic for reasons that "
     "have nothing to do with adoption."),
    ("06", "The data moves after you have exported it",
     "Activity typically appears within 48 hours and the previous three days are "
     "re-validated and backfilled on a rolling basis. Stamp every output with the "
     "refresh date and the window."),
    ("07", "A blank last-activity date is not always inactivity",
     "If a user used Copilot within 24 hours of licence assignment and never "
     "again, the last-activity date can be empty. Treat blank-with-nonzero-prompts "
     "as a data artefact, not as a user."),
    ("08", "Anonymized rows cannot be joined, and deleted users vanish",
     "With concealment on -- the default -- the principal-name column is a hash. "
     "Usage data for a deleted account is removed within 30 days, so totals will "
     "not reconcile across a period covering a departure."),
    ("09", "The audit log is not a usage report",
     "Microsoft states that Purview audit data is not intended as a basis for "
     "usage reporting and will not reconcile with the official reports. Use the "
     "usage report or the Copilot Dashboard."),
    ("10", "Zero is not plottable, and it is not divisible either",
     "A log depth axis cannot render zero prompts, and prompts-per-active-day "
     "divides by a column that is legitimately zero for much of the population. "
     "This workbook plots zero-prompt users at 1 and counts them on Setup and on "
     "the Chart caption. Say the excluded count out loud."),
]

REPORTS = [
    ("Microsoft Copilot usage report - user-level table",
     "Microsoft 365 admin center - admin.microsoft.com",
     "Reports > Usage > Microsoft Copilot > Copilot > the Usage tab. "
     "If Reports is not in the navigation menu, choose Show all first.",
     "7 / 28 / 90 / 180 days. Take 180 for a full read, or 90 if you are working "
     "at small-tenant thresholds. Do not use 7 or 28.",
     "Reports Reader, or any of: Usage Summary Reports Reader, AI Administrator, "
     "Exchange / SharePoint / Teams / Teams Communications Administrator, "
     "User Experience Success Manager, Global Administrator.",
     "Scroll to the user-level table, choose Choose columns and switch on "
     "everything, then use the Export on the table. The ellipsis Export on each "
     "chart gives you the chart's aggregate, not the user detail."),
    ("Microsoft Copilot agents usage report",
     "Microsoft 365 admin center - admin.microsoft.com",
     "Reports > Usage > Microsoft Copilot > Agents",
     "7 or 30 days only. Not 90, not 180.",
     "Same roles as the usage report. Names are concealed here too.",
     "Three tables: user details, agent details, and a user-and-agent pair view. "
     "The pair view tells you whether your champions have ever been given an agent."),
    ("Microsoft Copilot readiness report",
     "Microsoft 365 admin center - admin.microsoft.com",
     "Reports > Usage > Microsoft Copilot > Readiness",
     "Snapshot of eligibility and assignment rather than a rolling window.",
     "Same roles as the usage report.",
     "Who is eligible, who is licensed, where the gaps are. Use it when the tenant "
     "is not far enough along to have a distribution worth plotting."),
    ("Microsoft Copilot Dashboard - Viva Insights",
     "Microsoft Viva Insights. Quickest route in is the Recommendations card "
     "inside the Copilot usage report.",
     "Viva Insights > Copilot Dashboard",
     "Rolling windows set by the dashboard. Processing starts once at least one "
     "Copilot licence is assigned and takes up to seven days.",
     "Available with one or more Copilot licences. No paid Viva Insights licence "
     "required to view.",
     "Nothing, for this method. Read it to know what your customer has already "
     "been shown. Agent insights, benchmarks, group-level metrics, intelligent "
     "summaries and delegation are switched off below fifty Copilot licences."),
    ("Microsoft 365 usage reports (overview)",
     "Microsoft 365 admin center - admin.microsoft.com",
     "Reports > Usage",
     "7 / 30 / 90 / 180 days depending on the report.",
     "Reports Reader is sufficient.",
     "The landing page the Copilot report sits behind. Also where the concealment "
     "behaviour that applies to every report is documented."),
]

SOURCES = [
    ("Microsoft Copilot usage report",
     "https://learn.microsoft.com/en-us/microsoft-365/admin/activity-reports/microsoft-365-copilot-usage"),
    ("Microsoft 365 admin center usage reports overview",
     "https://learn.microsoft.com/en-us/microsoft-365/admin/activity-reports/activity-reports"),
    ("Graph API: getMicrosoft365CopilotUsageUserDetail",
     "https://learn.microsoft.com/en-us/microsoft-365/copilot/extensibility/api/admin-settings/reports/copilotreportroot-getmicrosoft365copilotusageuserdetail"),
    ("Copilot Agents usage report (preview)",
     "https://learn.microsoft.com/en-us/microsoft-365/admin/activity-reports/microsoft-365-copilot-agents-new"),
    ("Copilot Agent usage report (deprecated)",
     "https://learn.microsoft.com/en-us/microsoft-365/admin/activity-reports/microsoft-365-copilot-agents"),
    ("Copilot Dashboard in Viva Insights",
     "https://learn.microsoft.com/en-us/viva/insights/org-team-insights/copilot-dashboard"),
    ("Microsoft Copilot readiness report",
     "https://learn.microsoft.com/en-us/microsoft-365/admin/activity-reports/microsoft-365-copilot-readiness"),
    ("Call Microsoft Graph from a Cloud Solution Provider application",
     "https://learn.microsoft.com/en-us/graph/auth-cloudsolutionprovider"),
    ("Granular delegated admin privileges (GDAP) introduction",
     "https://learn.microsoft.com/en-us/partner-center/customers/gdap-introduction"),
    ("Authorization for APIs to read Microsoft 365 usage reports",
     "https://learn.microsoft.com/en-us/graph/reportroot-authorization"),
]

# --------------------------------------------------------------------------
# Small styling helpers
# --------------------------------------------------------------------------

thin = Side(style="thin", color=RULE)


def put(ws, ref, value, *, bold=False, size=10, color=None, italic=False,
        fill=None, wrap=False, fmt=None, align=None, valign="top", border=False,
        indent=0):
    """Write one cell with Arial styling and return it."""
    c = ws[ref]
    c.value = value
    c.font = Font(name=ARIAL, size=size, bold=bold, italic=italic,
                  color=color or "FF000000")
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(wrap_text=wrap, vertical=valign,
                            horizontal=align, indent=indent)
    if fmt:
        c.number_format = fmt
    if border:
        c.border = Border(bottom=thin)
    return c


def title(ws, ref, text):
    put(ws, ref, text, bold=True, size=14, color=F_NAVY)


def section(ws, ref, text):
    put(ws, ref, text, bold=True, size=11, color="FF" + TEAL)


def colhead(ws, row, first_col, labels, widths=None):
    for i, label in enumerate(labels):
        ref = f"{get_column_letter(first_col + i)}{row}"
        put(ws, ref, label, bold=True, size=9, color="FFFFFFFF", fill=NAVY,
            wrap=True, valign="center")
    if widths:
        for i, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(first_col + i)].width = w


def widths(ws, spec):
    for letter, w in spec.items():
        ws.column_dimensions[letter].width = w


# --------------------------------------------------------------------------
# Sheet builders
# --------------------------------------------------------------------------

def build_start_here(ws):
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 2.5, "B": 30, "C": 96})

    rows = [
        ("T", "Copilot Adoption Audit - workbook", None),
        ("N", "Paste one Microsoft Copilot usage report export into the Paste Export "
              "tab. Everything else computes itself.", None),
        ("S", "Where the data comes from", None),
        ("L", "Report", "Microsoft Copilot usage report - the user-level table."),
        ("L", "Portal", "Microsoft 365 admin center - admin.microsoft.com"),
        ("L", "Navigation", "Reports > Usage > Microsoft Copilot > Copilot > the Usage tab. "
                            "If Reports is not in the navigation menu, choose Show all first."),
        ("L", "Period", "7 / 28 / 90 / 180 days. Take 180 for a full read, or 90 if you are "
                        "working at small-tenant thresholds. Do not use 7 or 28 - they show "
                        "you noise and call it a trend."),
        ("L", "Role", "Reports Reader is sufficient. Not Global Administrator."),
        ("L", "Before exporting", "Scroll to the user-level table and choose Choose columns, "
                                  "then switch on everything. The default column set is not enough."),
        ("L", "What you take", "The Export button on the user-level table. One CSV, one row "
                               "per licensed user."),
        ("W", "The wrong Export", "The ellipsis menu on each individual chart also offers an "
                                  "Export. That gives you the chart's aggregate, not the user "
                                  "detail, and it is the mistake almost everyone makes the "
                                  "first time. You want the Export on the table."),
        ("S", "How to use this workbook", None),
        ("L", "1", "Open the CSV. Select everything including the header row, and copy."),
        ("L", "2", "Paste at cell A1 of the Paste Export tab - headers included. Delete the "
                   "three example rows first; they are there to show the expected shape."),
        ("L", "3", "Fill in the four yellow cells on Setup: customer name, report period in "
                   "days, refresh date, and a mode override if you want one."),
        ("L", "4", "Check the column-detection block on Setup. Every line should read FOUND."),
        ("L", "5", "Read Segments. That is the deliverable. Analysis is the per-user working, "
                   "Chart is the scatter, Book Ranking is the portfolio view."),
        ("S", "The two modes", None),
        ("L", "LARGE (60+ rows)", "Thresholds are this tenant's own medians - median active "
                                  "days and median prompts. The scatter chart is meaningful."),
        ("L", "SMALL (under 60)", "Thresholds are fixed: 9 active days, and 15 prompts per "
                                  "week across the period. A median computed on thirty rows "
                                  "moves when one person takes annual leave. Do not use the "
                                  "chart in this mode - read the ranked list on Analysis."),
        ("L", "Override", "Set Setup C7 to SMALL or LARGE to force a mode. Leave it blank to "
                          "let the row count decide. Whichever mode produced the numbers is "
                          "stamped on the Segments subtitle - it will be the first question "
                          "anyone asks six months later."),
        ("S", "Three rules this workbook enforces for you", None),
        ("L", "n<5 suppression", "No percentage, median or rate is reported for any band with "
                                 "fewer than five people in it. Those cells read "
                                 "'n<5 not reported'. The raw count is still shown."),
        ("L", "Zero-guarding", "Prompts per active day divides by a column that is "
                               "legitimately zero, and a log axis cannot plot zero prompts. "
                               "Rates are wrapped, plot depth is floored at 1, and the "
                               "excluded count is stated on Setup and on the chart caption."),
        ("L", "Correlational only", "This measures where the tool has not been made useful "
                                    "yet. It does not measure effort, contribution, "
                                    "productivity or performance. Say that in the room before "
                                    "the table goes on the screen, every time."),
        ("S", "Legend", None),
        ("Y", "Yellow fill", "Cells you edit. Everything else is a formula - do not overtype it."),
        ("B", "Blue text", "A typed value rather than a calculation."),
        ("G", "Green text", "A link to a result on another sheet in this workbook."),
        ("L", "Capacity", f"{CAPACITY:,} users. Every formula range runs to row {LAST_ROW}. "
                          "To extend, select the last row of Analysis and Calc, fill down as "
                          "far as you need, and widen the Setup, Segments and Chart ranges to "
                          "match. Twenty tenants of a few hundred seats each fit as they are."),
        ("W", "Before you paste a customer's file anywhere",
              "This is employee data. Get written permission, leave Microsoft's "
              "anonymization on, and decide where the file lives and when it is destroyed "
              "before the first one lands."),
    ]

    r = 2
    for kind, label, value in rows:
        if kind == "T":
            title(ws, f"B{r}", label)
            r += 2
            continue
        if kind == "N":
            put(ws, f"B{r}", label, size=10, italic=True, color=F_GREY, wrap=True)
            ws.merge_cells(f"B{r}:C{r}")
            r += 2
            continue
        if kind == "S":
            section(ws, f"B{r}", label)
            r += 1
            continue
        if kind == "W":
            put(ws, f"B{r}", label, bold=True, size=10, color=F_RED, wrap=True)
            put(ws, f"C{r}", value, size=10, color=F_RED, wrap=True)
        elif kind == "Y":
            put(ws, f"B{r}", label, bold=True, size=10, fill=YELLOW, wrap=True)
            put(ws, f"C{r}", value, size=10, wrap=True)
        elif kind == "B":
            put(ws, f"B{r}", label, bold=True, size=10, color=F_BLUE, wrap=True)
            put(ws, f"C{r}", value, size=10, wrap=True)
        elif kind == "G":
            put(ws, f"B{r}", label, bold=True, size=10, color=F_GREEN, wrap=True)
            put(ws, f"C{r}", value, size=10, wrap=True)
        else:
            put(ws, f"B{r}", label, bold=True, size=10, color=F_NAVY, wrap=True)
            put(ws, f"C{r}", value, size=10, wrap=True)
        ws.row_dimensions[r].height = max(14, 12 * (1 + len(value or "") // 105))
        r += 1


def build_paste_export(ws):
    for i, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(name=ARIAL, size=9, bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = 21

    for r, row in enumerate(EXAMPLES, start=2):
        for i, v in enumerate(row, start=1):
            c = ws.cell(row=r, column=i, value=(v if v != "" else None))
            c.font = Font(name=ARIAL, size=10, color=F_BLUE)
            c.fill = PatternFill("solid", fgColor=YELLOW)

    # The note must not sit in column A: 'Rows detected' on Setup is a COUNTA over
    # 'Paste Export'!A2:A1001 and would count it as a user.
    put(ws, "C6", "The three rows above are EXAMPLES ONLY - one heavy user, one middling, "
                  "one drifting. Delete them before you paste a real export. Paste at A1, "
                  "headers included.",
        size=10, bold=True, color=F_RED)
    # Deliberately not merged: Excel refuses to paste over a merged cell, and this
    # one sits inside the paste area. The text overflows the empty cells instead.
    ws.row_dimensions[6].height = 30
    ws.row_dimensions[1].height = 46

    ws["A1"].comment = Comment(
        "Column order does not matter.\n\n"
        "Rows 2 to 4 are examples only - one heavy user, one middling, one "
        "drifting. Delete them before you paste a real export.\n\n"
        "This workbook finds every column it needs by matching the header text in "
        "row 1, so it does not matter which columns the admin switched on or what "
        "order the export produced them in. Paste the export exactly as it came, "
        "headers included, starting at A1.\n\n"
        "Check the column-detection block on Setup after pasting: every line "
        "should read FOUND.",
        "Copilot Adoption Audit", width=380, height=210)

    ws.freeze_panes = "A2"


def build_setup(ws):
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 2.5, "B": 34, "C": 24, "D": 3, "E": 30, "F": 46, "G": 10, "H": 12})

    title(ws, "B2", "Setup")
    put(ws, "B3", "Fill in the yellow cells. Everything else is a formula.",
        size=10, italic=True, color=F_GREY)

    section(ws, "B5", "1 - Inputs")
    inputs = [
        ("B6", "Customer name", "C6", "Example Customer Ltd", None),
        ("B7", "Report period (days)", "C7", 90, "0"),
        ("B8", "Report refresh date (UTC)", "C8", "2026-08-26", None),
        ("B9", "Mode override (SMALL / LARGE / blank)", "C9", None, None),
    ]
    for lref, label, vref, val, fmt in inputs:
        put(ws, lref, label, size=10, bold=True, color=F_NAVY)
        put(ws, vref, val, size=10, color=F_BLUE, fill=YELLOW, fmt=fmt)
    put(ws, "E6", "The name that appears on the Segments subtitle.", size=9, color=F_GREY)
    put(ws, "E7", "7, 28, 90 or 180. Drives the SMALL-mode depth line.", size=9, color=F_GREY)
    put(ws, "E8", "Shown at the top of the usage report. Type it as text if you prefer.",
        size=9, color=F_GREY)
    put(ws, "E9", "Leave blank to let the row count decide.", size=9, color=F_GREY)

    section(ws, "B11", "2 - Detected from the pasted data")
    put(ws, "B12", "Rows detected", size=10, bold=True, color=F_NAVY)
    put(ws, "C12", f"=COUNTA('Paste Export'!A2:A{LAST_ROW})", size=10, fmt="0")
    put(ws, "B13", "Automatic mode", size=10, bold=True, color=F_NAVY)
    put(ws, "C13", '=IF(C12>=60,"LARGE","SMALL")', size=10)
    put(ws, "B14", "Mode in use", size=10, bold=True, color=F_NAVY)
    put(ws, "C14", '=IF(C9="",C13,UPPER(C9))', size=10, bold=True)
    put(ws, "E12", f"COUNTA over the first column of the paste area. Capacity {CAPACITY:,} rows.",
        size=9, color=F_GREY)
    put(ws, "E13", "Sixty rows is where a median stops moving when one person takes leave.",
        size=9, color=F_GREY)
    put(ws, "E14", "This is stamped on the Segments subtitle so the numbers are traceable.",
        size=9, color=F_GREY)

    section(ws, "B16", "3 - Column detection (by header text, never by position)")
    put(ws, "B17", "Each line matches the exact Microsoft header string against row 1 of "
                   "Paste Export. Nothing in this workbook refers to a column letter of the "
                   "pasted data.", size=9, italic=True, color=F_GREY, wrap=True)
    ws.merge_cells("B17:F17")
    ws.row_dimensions[17].height = 26

    colhead(ws, 18, 2, ["Field", "", "", "Header text matched", "Column", "Status"])
    ws["B18"].alignment = Alignment(wrap_text=True, vertical="center")
    for i, (label, header, _name) in enumerate(DETECT):
        r = 19 + i
        put(ws, f"B{r}", label, size=10, color=F_NAVY, bold=True)
        put(ws, f"E{r}", header, size=9, color=F_GREY, wrap=True)
        put(ws, f"F{r}", f'=IFERROR(MATCH($E{r},{HDR},0),0)', size=10, fmt="0")
        put(ws, f"G{r}", f'=IF($F{r}=0,"NOT FOUND","FOUND")', size=9, bold=True)

    first, last = 19, 19 + len(DETECT) - 1
    put(ws, f"B{last + 2}",
        f'=IF(COUNTIF($F${first}:$F${last},0)=0,'
        f'"All {len(DETECT)} columns found.",'
        f'"MISSING COLUMNS - "&COUNTIF($F${first}:$F${last},0)&'
        f'" of {len(DETECT)}. Check you exported the user table with all columns switched on, '
        f'and that this is a v2 report. Version 1 has no prompt or active-day columns at all.")',
        size=10, bold=True, color=F_RED, wrap=True)
    ws.merge_cells(f"B{last + 2}:F{last + 2}")
    ws.row_dimensions[last + 2].height = 30

    tr = last + 4
    section(ws, f"B{tr}", "4 - Thresholds")
    put(ws, f"B{tr + 1}", "Frequency line - active days", size=10, bold=True, color=F_NAVY)
    put(ws, f"C{tr + 1}",
        f'=IF(run_mode="SMALL", 9, MEDIAN(INDEX({PASTE},0,col_days)))', size=10, fmt="0.0")
    put(ws, f"B{tr + 2}", "Depth line - prompts", size=10, bold=True, color=F_NAVY)
    put(ws, f"C{tr + 2}",
        f'=IF(run_mode="SMALL", ROUND(15*(period_days/7),0), MEDIAN(INDEX({PASTE},0,col_prompts)))',
        size=10, fmt="0.0")
    put(ws, f"E{tr + 1}",
        '=IF(run_mode="SMALL","Fixed: active in 9 of the past 12 weeks.",'
        '"This tenant\'s own median active days.")', size=9, color=F_GREY, wrap=True)
    put(ws, f"E{tr + 2}",
        '=IF(run_mode="SMALL","Fixed: 15 prompts per week across "&period_days&" days.",'
        '"This tenant\'s own median prompts.")', size=9, color=F_GREY, wrap=True)
    put(ws, f"B{tr + 3}",
        "Band mapping: high frequency + high depth = Embedded. High frequency + low depth = "
        "Loyal but shallow. Low frequency + high depth = Project-driven. Low + low = Drifting.",
        size=9, italic=True, color=F_GREY, wrap=True)
    ws.merge_cells(f"B{tr + 3}:F{tr + 3}")
    ws.row_dimensions[tr + 3].height = 26

    dr = tr + 5
    section(ws, f"B{dr}", "5 - Disclosure counters (say these out loud)")
    put(ws, f"B{dr + 1}", "Never-active users", size=10, bold=True, color=F_NAVY)
    put(ws, f"C{dr + 1}", f"=SUM(Analysis!$L$2:$L${LAST_ROW})", size=10, fmt="0")
    put(ws, f"E{dr + 1}", "Zero prompts and zero active days. The most defensible number here.",
        size=9, color=F_GREY, wrap=True)
    put(ws, f"B{dr + 2}", "Excluded from the chart (zero prompts)", size=10, bold=True, color=F_NAVY)
    put(ws, f"C{dr + 2}", f'=COUNTIF(Analysis!$D$2:$D${LAST_ROW},0)', size=10, fmt="0")
    put(ws, f"E{dr + 2}", "A log axis cannot render zero. These are plotted at 1 - state the "
                          "count on the slide.", size=9, color=F_GREY, wrap=True)
    put(ws, f"B{dr + 3}", "Surfaces counted per user", size=10, bold=True, color=F_NAVY)
    put(ws, f"C{dr + 3}", "8", size=10)
    put(ws, f"E{dr + 3}", "Teams, Word, Excel, PowerPoint, Outlook, OneNote, Loop and the "
                          "Microsoft 365 App. Edge and the two chat entry points are chat "
                          "surfaces, not application surfaces, and are excluded from the count.",
        size=9, color=F_GREY, wrap=True)
    ws.row_dimensions[dr + 3].height = 28

    return {"tr": tr, "dr": dr}


def build_analysis(ws):
    heads = ["#", "User", "Active days", "Prompts", "Web prompts",
             "Prompts / active day", "Web share of prompts", "Surfaces touched",
             "Band", "Drifting split", "Plot depth", "Never active"]
    colhead(ws, 1, 1, heads,
            widths=[6, 22, 11, 10, 11, 12, 12, 11, 17, 17, 10, 10])
    ws.row_dimensions[1].height = 40
    ws.freeze_panes = "B2"

    for r in range(2, LAST_ROW + 1):
        g = f'IF($A{r}>nrows,"",'
        surf = "+".join(
            f'IF({n}=0,0,--(LEN(INDEX({PASTE},$A{r},{n}))>0))' for n in SURF_NAMES)
        cells = {
            "A": "=ROW()-1",
            "B": f'={g}IF(col_user=0,"",INDEX({PASTE},$A{r},col_user)))',
            "C": f'={g}IF(col_days=0,0,N(INDEX({PASTE},$A{r},col_days))))',
            "D": f'={g}IF(col_prompts=0,0,N(INDEX({PASTE},$A{r},col_prompts))))',
            "E": f'={g}IF(col_web=0,0,N(INDEX({PASTE},$A{r},col_web))))',
            "F": f'={g}IFERROR($D{r}/$C{r},0))',
            "G": f'={g}IFERROR($E{r}/$D{r},""))',
            "H": f"={g}{surf})",
            "I": (f'={g}IF($C{r}>=freq_line,IF($D{r}>=depth_line,"Embedded",'
                  f'"Loyal but shallow"),IF($D{r}>=depth_line,"Project-driven","Drifting")))'),
            "J": (f'={g}IF($I{r}="Drifting",IF($H{r}>=3,"tried and stopped",'
                  f'"never started"),""))'),
            "K": f"={g}MAX($D{r},1))",
            "L": f"={g}IF(AND($C{r}=0,$D{r}=0),1,0))",
        }
        for col, formula in cells.items():
            c = ws[f"{col}{r}"]
            c.value = formula
            c.font = Font(name=ARIAL, size=10)
        ws[f"A{r}"].number_format = "0"
        for col in "CDEHKL":
            ws[f"{col}{r}"].number_format = "0"
        ws[f"F{r}"].number_format = "0.0"
        ws[f"G{r}"].number_format = "0.0%"


def build_calc(ws):
    """Twelve-plus helper columns: one per (metric, band). MEDIAN ignores blanks."""
    metrics = [("Prompts", "D"), ("Active days", "C"),
               ("Prompts / active day", "F"), ("Surfaces", "H")]
    idx = 1
    layout = {}
    for mname, mcol in metrics:
        for band in BANDS:
            letter = get_column_letter(idx)
            layout[(mname, band)] = letter
            c = ws.cell(row=1, column=idx, value=f"{band} - {mname}")
            c.font = Font(name=ARIAL, size=8, bold=True, color="FFFFFFFF")
            c.fill = PatternFill("solid", fgColor=NAVY)
            c.alignment = Alignment(wrap_text=True, vertical="center")
            ws.column_dimensions[letter].width = 13
            for r in range(2, LAST_ROW + 1):
                cell = ws.cell(
                    row=r, column=idx,
                    value=(f'=IF(Analysis!$I{r}="","",'
                           f'IF(Analysis!$I{r}="{band}",Analysis!${mcol}{r},""))'))
                cell.font = Font(name=ARIAL, size=9)
            idx += 1
    ws.row_dimensions[1].height = 34
    ws.freeze_panes = "A2"
    return layout


def build_segments(ws, calc):
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 20, "B": 34, "C": 9, "D": 13, "E": 15, "F": 13, "G": 13,
                "H": 13, "I": 11, "J": 62})

    title(ws, "A1", "Segments")
    put(ws, "A2",
        '=cust_name&" - "&nrows&" users - "&period_days&"-day window - refreshed "&'
        'IF(refresh_date="","date not stated",TEXT(refresh_date,"d mmm yyyy"))&'
        '" - "&run_mode&" mode"',
        size=11, italic=True, color=F_GREY)
    put(ws, "A3",
        "Relationships shown are correlational. This measures where the tool has not been "
        "made useful yet - not effort, contribution or value.",
        size=10, bold=True, color=F_RED, wrap=True)
    ws.merge_cells("A3:J3")
    ws.row_dimensions[3].height = 18

    heads = ["Band", "Threshold rule", "Users", "% of users", "% of all prompts",
             "Median prompts", "Median active days", "Median prompts / active day",
             "Median surfaces", "What it means"]
    colhead(ws, 5, 1, heads)
    ws.row_dimensions[5].height = 42

    for i, band in enumerate(BANDS):
        r = 6 + i
        put(ws, f"A{r}", band, size=10, bold=True, color=F_NAVY, valign="center")
        put(ws, f"B{r}", BAND_RULE[band], size=9, color=F_GREY, wrap=True, valign="center")
        put(ws, f"C{r}", f'=COUNTIF(Analysis!$I$2:$I${LAST_ROW},$A{r})',
            size=10, fmt="0", valign="center")
        put(ws, f"D{r}",
            f'=IF($C{r}<5,"n<5 not reported",IFERROR($C{r}/nrows,0))',
            size=10, fmt="0.0%", valign="center")
        put(ws, f"E{r}",
            f'=IF($C{r}<5,"n<5 not reported",'
            f'IFERROR(SUMIF(Analysis!$I$2:$I${LAST_ROW},$A{r},Analysis!$D$2:$D${LAST_ROW})'
            f'/SUM(Analysis!$D$2:$D${LAST_ROW}),0))',
            size=10, fmt="0.0%", valign="center")
        for col, metric, fmt in (("F", "Prompts", "0"),
                                 ("G", "Active days", "0"),
                                 ("H", "Prompts / active day", "0.0"),
                                 ("I", "Surfaces", "0")):
            cl = calc[(metric, band)]
            put(ws, f"{col}{r}",
                f'=IF($C{r}<5,"n<5 not reported",MEDIAN(Calc!${cl}$2:${cl}${LAST_ROW}))',
                size=10, fmt=fmt, valign="center")
        put(ws, f"J{r}", BAND_MEANING[band], size=9, wrap=True, valign="center")
        ws.row_dimensions[r].height = 56

    put(ws, "A10", "All bands", size=10, bold=True, color=F_NAVY)
    put(ws, "C10", "=SUM(C6:C9)", size=10, bold=True, fmt="0")
    put(ws, "B10", "Should equal the rows detected on Setup.", size=9, color=F_GREY)
    put(ws, "D10", '=IF(nrows=0,"",SUM(C6:C9)/nrows)', size=10, bold=True, fmt="0.0%")
    put(ws, "E10", f'=IF(SUM(Analysis!$D$2:$D${LAST_ROW})=0,"",1)', size=10, bold=True, fmt="0.0%")

    section(ws, "A12", "The finding")
    put(ws, "A13",
        '=IF($C$6<5,'
        '"Fewer than five Embedded users - do not quote a share. Report the count ("&$C$6&'
        '") and the never-active number instead, and say plainly that the sample is too '
        'small for a percentage.",'
        '"Embedded users are "&TEXT($D$6,"0.0%")&" of the licensed population and generate "&'
        'TEXT($E$6,"0.0%")&" of all prompts. That gap is the finding: it is where the return '
        'is concentrated, and how thin it is.")',
        size=10, wrap=True)
    ws.merge_cells("A13:J13")
    ws.row_dimensions[13].height = 30
    put(ws, "A14",
        '=IF($C$9<5,'
        '"Fewer than five Drifting users - report the count ("&$C$9&") only.",'
        '"Drifting users are "&TEXT($D$9,"0.0%")&" of the population and generate "&'
        'TEXT($E$9,"0.0%")&" of all prompts. Exposure happened; a reason to come back did '
        'not. Split the band below before acting on it.")',
        size=10, wrap=True)
    ws.merge_cells("A14:J14")
    ws.row_dimensions[14].height = 30

    section(ws, "A16", "Licence exposure - four numbers, and they are never merged")
    exposure = [
        ("Never active (zero prompts and zero active days)",
         f"=SUM(Analysis!$L$2:$L${LAST_ROW})"),
        ("Zero prompts (excluded from the chart)",
         f'=COUNTIF(Analysis!$D$2:$D${LAST_ROW},0)'),
        ("Drifting - never started (under 3 surfaces touched)",
         f'=COUNTIF(Analysis!$J$2:$J${LAST_ROW},"never started")'),
        ("Drifting - tried and stopped (3 or more surfaces touched)",
         f'=COUNTIF(Analysis!$J$2:$J${LAST_ROW},"tried and stopped")'),
    ]
    for i, (label, formula) in enumerate(exposure):
        r = 17 + i
        put(ws, f"A{r}", label, size=10, color=F_NAVY, wrap=True)
        ws.merge_cells(f"A{r}:B{r}")
        put(ws, f"C{r}", formula, size=10, bold=True, fmt="0")
    put(ws, "A21",
        "These four overlap and they are not a wasted-licence total. Do not add them "
        "together and do not put a single number on a slide. Reconcile the row count "
        "against currently assigned licences under Billing > Licenses first: the table "
        "shows everyone who held a Copilot licence at any point in the window, including "
        "people whose licence was removed.",
        size=9, color=F_RED, wrap=True)
    ws.merge_cells("A21:J21")
    ws.row_dimensions[21].height = 30

    section(ws, "A23", "The realistic ceiling - prompts per active day, in this tenant")
    # PERCENTILE over an empty population returns #NUM!, so the zero-row case is
    # guarded rather than left to error before anyone has pasted anything.
    ceiling = [
        ("Median", f'=IF(nrows=0,0,PERCENTILE(Analysis!$F$2:$F${LAST_ROW},0.5))'),
        ("75th percentile", f'=IF(nrows=0,0,PERCENTILE(Analysis!$F$2:$F${LAST_ROW},0.75))'),
        ("95th percentile", f'=IF(nrows=0,0,PERCENTILE(Analysis!$F$2:$F${LAST_ROW},0.95))'),
        ("Maximum", f"=MAX(Analysis!$F$2:$F${LAST_ROW})"),
    ]
    for i, (label, formula) in enumerate(ceiling):
        r = 24 + i
        put(ws, f"A{r}", label, size=10, color=F_NAVY)
        put(ws, f"C{r}", formula, size=10, bold=True, fmt="0.0")
    put(ws, "A28",
        '=IF(nrows<60,"Under 60 users - quote the median and the maximum only. A 75th or '
        '95th percentile computed on this few people is one person, and someone in the room '
        'will work out who.","Use these to replace an invented adoption target with a '
        'defensible one drawn from this tenant\'s own distribution.")',
        size=9, color=F_RED, wrap=True)
    ws.merge_cells("A28:J28")
    ws.row_dimensions[28].height = 28


def build_chart(ws):
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 18, "B": 18, "C": 18, "D": 18, "E": 18, "F": 18, "G": 18, "H": 18})

    title(ws, "A1", "Frequency against depth")
    put(ws, "A2",
        '=IF(run_mode="SMALL","SMALL MODE - do not use this chart. Read the ranked list '
        'on Analysis instead.","")',
        size=11, bold=True, color=F_RED)
    put(ws, "A3",
        f'="Excluded from the plot: "&COUNTIF(Analysis!$D$2:$D${LAST_ROW},0)&'
        '" users with zero prompts (a logarithmic axis cannot render zero; they are '
        'plotted at 1). Report refreshed "&IF(refresh_date="","date not stated",'
        'TEXT(refresh_date,"d mmm yyyy"))&", "&period_days&"-day window, "&nrows&'
        '" users. Relationships shown are correlational - this is a map of unrealised '
        'opportunity, not a scorecard of outcome."',
        size=9, color=F_GREY, wrap=True)
    ws.merge_cells("A3:H3")
    ws.row_dimensions[3].height = 42

    chart = ScatterChart()
    chart.title = "Chat frequency against depth"
    chart.style = 13
    chart.x_axis.title = "Active days (chat frequency)"
    chart.y_axis.title = "Prompts (plot depth, log scale)"
    chart.y_axis.scaling.logBase = 10
    chart.y_axis.scaling.min = 1
    chart.x_axis.majorGridlines = None
    chart.height = 14
    chart.width = 24
    chart.legend = None

    xref = Reference(ws.parent["Analysis"], min_col=3, min_row=2, max_row=LAST_ROW)
    yref = Reference(ws.parent["Analysis"], min_col=11, min_row=2, max_row=LAST_ROW)
    s = Series(yref, xref, title="Users")
    s.marker = Marker(symbol="circle", size=4)
    s.graphicalProperties.line.noFill = True
    chart.series.append(s)
    ws.add_chart(chart, "A5")

    put(ws, "A34",
        "Every point is one user. Horizontal position is how often they used Copilot Chat; "
        "vertical position is how much. The two threshold lines are on Setup - change them "
        "there and the whole population re-sorts.",
        size=9, italic=True, color=F_GREY, wrap=True)
    ws.merge_cells("A34:H34")


def build_book_ranking(ws):
    ws.sheet_view.showGridLines = False
    title(ws, "A1", "Book ranking - one row per Copilot customer")
    put(ws, "A2",
        "Row 5 reads this workbook's own result in green. Rows 6 and below are yellow: "
        "one row per customer, filled in from that customer's copy of this workbook. Sort "
        "by never-active share descending and filter to renewals inside two quarters, and "
        "you have your priority list.",
        size=10, italic=True, color=F_GREY, wrap=True)
    ws.merge_cells("A2:L2")
    ws.row_dimensions[2].height = 28

    heads = ["Customer", "Licensed users", "Never active", "Never active %",
             "Drifting %", "Embedded %", "Embedded with no agent use",
             "Web-only users", "Months to renewal", "Next play", "Owner", "Date agreed"]
    colhead(ws, 4, 1, heads,
            widths=[26, 12, 11, 11, 11, 11, 15, 12, 12, 26, 16, 13])
    ws.row_dimensions[4].height = 46

    live = {
        "A": "=cust_name",
        "B": "=nrows",
        "C": f"=SUM(Analysis!$L$2:$L${LAST_ROW})",
        "D": "=IFERROR(C5/B5,0)",
        "E": "=Segments!$D$9",
        "F": "=Segments!$D$6",
        "H": f'=COUNTIF(Analysis!$G$2:$G${LAST_ROW},">=0.9")',
    }
    fmts = {"B": "0", "C": "0", "D": "0.0%", "E": "0.0%", "F": "0.0%", "H": "0"}
    for col, formula in live.items():
        put(ws, f"{col}5", formula, size=10, color=F_GREEN, bold=True, fmt=fmts.get(col))
    for col in ("G", "I", "J", "K", "L"):
        put(ws, f"{col}5", None, size=10, color=F_BLUE, fill=YELLOW)
    put(ws, "N5",
        "G comes from the agents usage report, not from this export - the two reports do "
        "not share a window, so state both. I to L come from your own systems.",
        size=9, color=F_GREY, wrap=True)

    for r in range(6, 26):
        for col in "ABCDEFGHIJKL":
            put(ws, f"{col}{r}", None, size=10, color=F_BLUE, fill=YELLOW,
                fmt=fmts.get(col))

    put(ws, "A27",
        "Next play - one of four: use-case design by job role (Drifting), a second use "
        "case in thirty minutes (Loyal but shallow), workflow embedding plus a grounding "
        "conversation (Project-driven), agent access rather than agent authorship "
        "(Embedded). Assign an owner and a date or the exercise stays interesting and "
        "never becomes revenue.",
        size=9, italic=True, color=F_GREY, wrap=True)
    ws.merge_cells("A27:L27")
    ws.row_dimensions[27].height = 30
    put(ws, "A29",
        "Rows are only comparable if they were produced in the same mode. The mode is "
        "stamped on each workbook's Segments subtitle - record it if you mix tenant sizes.",
        size=9, color=F_RED, wrap=True)
    ws.merge_cells("A29:L29")


def build_reference(ws):
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 2.5, "B": 30, "C": 100})

    title(ws, "B2", "Reference")
    put(ws, "B3",
        "The offline twin of Sections 8, 9 and 13 of the Copilot Adoption Audit page. "
        "Microsoft's reporting surface changes frequently - re-verify column names, report "
        "paths and window lengths before you re-deliver.",
        size=10, italic=True, color=F_GREY, wrap=True)
    ws.merge_cells("B3:C3")
    ws.row_dimensions[3].height = 28

    r = 5
    section(ws, f"B{r}", "The five reports")
    r += 1
    for name, portal, nav, period, role, take in REPORTS:
        put(ws, f"B{r}", name, size=10, bold=True, color=F_NAVY, wrap=True)
        ws.merge_cells(f"B{r}:C{r}")
        r += 1
        for label, value in (("Portal", portal), ("Navigation", nav),
                             ("Period", period), ("Role", role), ("What you take", take)):
            put(ws, f"B{r}", label, size=9, bold=True, color=F_GREY)
            put(ws, f"C{r}", value, size=9, wrap=True)
            ws.row_dimensions[r].height = max(13, 11 * (1 + len(value) // 110))
            r += 1
        r += 1

    section(ws, f"B{r}", "The concealment setting")
    r += 1
    for label, value in (
        ("What it does", "Conceals user, group and site names in every Microsoft 365 usage "
                         "report. It is on by default, and every band, segment size and "
                         "chart in this method works perfectly against hashes."),
        ("Path", "Settings > Org Settings > Services > Reports > "
                 "\"Conceal user, group, and site names in all reports\""),
        ("Role", "Global Administrator - a materially bigger ask than Reports Reader, and "
                 "the reason to avoid needing it."),
        ("Scope", "Tenant-wide and retroactive across every usage report, including Graph, "
                  "Power BI and the Teams admin center. Not scoped to your export."),
        ("Audit", "Showing identifiable user information is a logged event in the Microsoft "
                  "Purview audit log. Someone will see it."),
        ("The better ask", "Have the customer's admin add a department or job-family column "
                           "on their side and hand you the file with identities already "
                           "stripped. It succeeds far more often and leaves nothing to "
                           "explain."),
    ):
        put(ws, f"B{r}", label, size=9, bold=True, color=F_GREY)
        put(ws, f"C{r}", value, size=9, wrap=True)
        ws.row_dimensions[r].height = max(13, 11 * (1 + len(value) // 110))
        r += 1
    r += 1

    section(ws, f"B{r}", "If you would rather script it than click it")
    r += 1
    put(ws, f"B{r}", "Graph call", size=9, bold=True, color=F_GREY)
    put(ws, f"C{r}",
        "GET https://graph.microsoft.com/v1.0/copilot/reports/"
        "getMicrosoft365CopilotUsageUserDetail(period='D180', version='v2')",
        size=9, wrap=True)
    ws.row_dimensions[r].height = 26
    r += 1
    put(ws, f"B{r}", "Permission", size=9, bold=True, color=F_GREY)
    put(ws, f"C{r}", "Reports.Read.All - delegated or application. The response is a CSV stream.",
        size=9, wrap=True)
    r += 1
    put(ws, f"B{r}", "The version trap", size=9, bold=True, color=F_RED)
    put(ws, f"C{r}",
        "Version 1 - the default if you omit the parameter - contains no prompt counts and "
        "no active days. Both axes of this method are missing from it. You must pass "
        "version='v2'. Period values differ by version too: v1 accepts D7 D30 D90 D180 ALL, "
        "v2 accepts D7 D28 D90 D180 ALL.",
        size=9, color=F_RED, wrap=True)
    ws.row_dimensions[r].height = 40
    r += 2

    section(ws, f"B{r}", "The ten traps, one line each")
    r += 1
    for num, head, line in TRAPS:
        put(ws, f"B{r}", f"{num}  {head}", size=9, bold=True, color=F_NAVY, wrap=True)
        put(ws, f"C{r}", line, size=9, wrap=True)
        ws.row_dimensions[r].height = max(24, 11 * (1 + len(line) // 110))
        r += 1
    r += 1
    put(ws, f"B{r}", "The meta-trap", size=9, bold=True, color=F_RED)
    put(ws, f"C{r}",
        "Nine of the ten make your numbers look worse than reality if you ignore them. An "
        "unexamined export systematically over-states failure - and the admin in the room "
        "who knows these definitions will dismantle you.",
        size=9, color=F_RED, wrap=True)
    ws.row_dimensions[r].height = 30
    r += 2

    section(ws, f"B{r}", "Sources")
    r += 1
    for name, url in SOURCES:
        put(ws, f"B{r}", name, size=9, bold=True, color=F_NAVY, wrap=True)
        put(ws, f"C{r}", url, size=9, color=F_GREY)
        r += 1


# --------------------------------------------------------------------------
# Assembly
# --------------------------------------------------------------------------

def build(path: Path) -> None:
    wb = Workbook()
    ws_start = wb.active
    ws_start.title = "Start Here"
    ws_paste = wb.create_sheet("Paste Export")
    ws_setup = wb.create_sheet("Setup")
    ws_analysis = wb.create_sheet("Analysis")
    ws_calc = wb.create_sheet("Calc")
    ws_segments = wb.create_sheet("Segments")
    ws_chart = wb.create_sheet("Chart")
    ws_book = wb.create_sheet("Book Ranking")
    ws_ref = wb.create_sheet("Reference")

    build_start_here(ws_start)
    build_paste_export(ws_paste)
    anchors = build_setup(ws_setup)

    # Defined names, so that no formula anywhere carries a bare Setup cell address.
    names = {
        "cust_name": "Setup!$C$6",
        "period_days": "Setup!$C$7",
        "refresh_date": "Setup!$C$8",
        "nrows": "Setup!$C$12",
        "run_mode": "Setup!$C$14",
        "freq_line": f"Setup!$C${anchors['tr'] + 1}",
        "depth_line": f"Setup!$C${anchors['tr'] + 2}",
    }
    for i, (_label, _header, name) in enumerate(DETECT):
        names[name] = f"Setup!$F${19 + i}"
    for name, ref in names.items():
        wb.defined_names.add(DefinedName(name, attr_text=ref))

    build_analysis(ws_analysis)
    calc = build_calc(ws_calc)
    build_segments(ws_segments, calc)
    build_chart(ws_chart)
    build_book_ranking(ws_book)
    build_reference(ws_ref)

    ws_calc.sheet_state = "hidden"
    wb.active = 0
    wb.save(path)


def main() -> int:
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else (
        Path(__file__).resolve().parent.parent / "copilot-adoption-audit-workbook.xlsx")
    build(out)
    print(f"wrote {out} ({out.stat().st_size:,} bytes)")
    print("now run: python scripts/recalc.py "
          f"{out.name} 120   # from the xlsx skill directory")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
